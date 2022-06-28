import urllib
from openpyxl import load_workbook,Workbook
import csv
from products.models import ProductSimple
from products.models import ProductBooleanEav,ProductCharEav,ProductIntEav,ProductDecimalEav,ProductTextEav,ProductUrlEav
from django.db.utils import DataError
from django.conf import settings
import os
from io import BytesIO
from PIL import Image
import urllib

 
        
class WorkBook():
    
    def load_workbook_from_path(self,path,isCSV=False):
        self.book = Workbook()
        if isCSV:
            csv_data=[]
            with open(path,"r",encoding="UTF-8") as fp:
                reader=csv.reader(fp,delimiter="\t")
                for row in reader:
                    csv_data.append(row)
            sheet = self.book.active
            for row in csv_data:
                sheet.append(row)
        else:
            self.book=load_workbook(path)

    # def load_workbook_from_url(self,url):
    #     file = urllib.request.urlopen(url).read()
    #     self.book=load_workbook(filename=file)

    # def load_workbook_from_stream(self,filestream):
    #     self.book=load_workbook(filename=filestream)
    #     filestream.close()
        
    def sheetnames(self):
        return self.book.sheetnames

    def close(self):
        self.book.close()

    def datasheet(self,sheetindex,cut=False,limit=None):
        ws=self.book.worksheets[sheetindex]
        wslist=[]
        i=0
        for row in ws.iter_rows():
            if limit is not None and i>limit:
                break
            i+=1
            stop=True
            ws_row=[]
            for cell in row:
                if cell.value is not None:
                    if cut:
                        ws_row.append(str(cell.value)[:80])
                    else:
                        ws_row.append(str(cell.value))
                    stop=False
                else:
                    ws_row.append("")
            if stop:
                return wslist
            else:
                wslist.append(ws_row)
        return wslist


    def importDatasheetToProducts(self,company,sheetindex,jump,columns,rows,marketplaces):
        ws=self.book.worksheets[sheetindex]
        wslist=[]
        messages={}
        updateType="N" # N,U,Overwrite
        j=0
        k=0
        try:
            for row in ws.iter_rows():
                k+=1
                messages[k]={}
                #salto le righe che non servono
                if jump>=0:
                    jump-=1
                    messages[k]["Riga"]="Saltata"
                    continue
                
                j+=1
                try:
                    if rows[j-1] is not True:
                        messages[k]["Riga"]="Riga da non elaborare"
                        continue
                except IndexError:
                    messages[k]["Riga"]="Saltata. Riga non valida"
                    continue
                
                
                i=0
                product_data={}
                product_data["eav_int"]={}
                product_data["eav_boolean"]={}
                product_data["eav_char"]={}
                product_data["eav_text"]={}
                product_data["eav_url"]={}
                product_data["eav_decimal"]={}
                for cell in row:
                    if str(i) in columns:
                        if(columns[str(i)]["name"] in ["sku","gtin","gtin_type"]):
                            product_data[columns[str(i)]["name"]]=cell.value
                        else:
                            if columns[str(i)]["type"]=="INT":
                                product_data["eav_int"][columns[str(i)]["name"]]=int(cell.value)
                            elif columns[str(i)]["type"]=="TEXT":
                                product_data["eav_text"][columns[str(i)]["name"]]=cell.value
                            elif columns[str(i)]["type"]=="BOOLEAN":
                                product_data["eav_boolean"][columns[str(i)]["name"]]=bool(cell.value)
                            elif columns[str(i)]["type"]=="DECIMAL":
                                product_data["eav_decimal"][columns[str(i)]["name"]]=cell.value
                            elif columns[str(i)]["type"]=="URL":
                                product_data["eav_url"][columns[str(i)]["name"]]=cell.value
                            elif columns[str(i)]["type"]=="CHAR":
                                product_data["eav_char"][columns[str(i)]["name"]]=str(cell.value)
                    i+=1
                try:
                    if product_data["sku"] in ["",None]:
                        messages[k]["Riga"]="Saltata. Devi inserire lo SKU"
                        continue
                except KeyError:
                    messages[k]["Riga"]="Saltata. Devi indicare la colonna degli SKU"
                    continue
                stroke=False
                product=None
                if updateType=="N": # Only new products
                    if ProductSimple.objects.filter(company=company,sku=product_data["sku"]).exists():
                        messages[k]["Riga"]="Saltata. Lo SKU esiste e la modalità di import è NUOVO INSERIMENTO"
                        continue
                    product=ProductSimple(company=company,sku=product_data["sku"])
                    try:
                        product.save()
                    except DataError as exc:
                        messages[k]["Riga"]="Saltata. "+str(exc)
                        continue
                elif updateType=="U": # Only existing products
                    try:
                        product=ProductSimple.objects.get(company=company,sku=product_data["sku"])
                    except:
                        messages[k]["Riga"]="Saltata. Lo SKU non esiste"
                        continue
                elif updateType=="O":
                    if ProductSimple.objects.filter(company=company,sku=product_data["sku"]).exists():
                        product=ProductSimple.objects.get(company=company,sku=product_data["sku"])
                    else:
                        product=ProductSimple(company=company,sku=product_data["sku"])
                        try:
                            product.save()
                        except DataError as exc:
                            messages[k]["Riga"]="Saltata. "+str(exc)
                            continue
                for key,value in product_data.items():
                    if(stroke):
                        break
                    if key == "gtin":
                        product.gtin=value
                    elif key == "gtin_type":
                        product.gtin_type=value
                    else:
                        
                        for marketplace in marketplaces.all():
                            if(stroke):
                                break
                            if key=="eav_int":
                                for kkey,vvalue in value.items():
                                    
                                    intProduct=ProductIntEav(sku=product_data["sku"],company=company,attribute=kkey,value=vvalue,marketplace=marketplace)
                                    try:
                                        intProduct.save()
                                        product.add(intProduct)
                                    except Exception as exc:
                                        messages[k][kkey]="Errore. "+str(exc)
                                        if updateType=="N":
                                            product.delete_eav(marketplace)
                                        stroke=True
                                        break
                            elif key=="eav_decimal":
                                for kkey,vvalue in value.items():
                                    decimalProduct=ProductDecimalEav(sku=product_data["sku"],company=company,attribute=kkey,value=vvalue,marketplace=marketplace)
                                    try:
                                        decimalProduct.save()
                                        product.decimal_eav.add(decimalProduct)
                                    except Exception as exc:
                                        messages[k][kkey]="Errore. "+str(exc)
                                        if updateType=="N":
                                            product.delete_eav(marketplace)
                                        stroke=True
                                        break
                            elif key=="eav_boolean":
                                for kkey,vvalue in value.items():
                                    booleanProduct=ProductBooleanEav(sku=product_data["sku"],company=company,attribute=kkey,value=vvalue,marketplace=marketplace)
                                    try:
                                        booleanProduct.save()
                                        product.boolean_eav.add(booleanProduct)
                                    except Exception as exc:
                                        messages[k][kkey]="Errore. "+str(exc)
                                        if updateType=="N":
                                            product.delete_eav(marketplace)
                                        stroke=True
                                        break
                            elif key=="eav_text":
                                for kkey,vvalue in value.items():
                                    textProduct=ProductTextEav(sku=product_data["sku"],company=company,attribute=kkey,value=vvalue,marketplace=marketplace)
                                    try:
                                        textProduct.save()
                                        product.text_eav.add(textProduct)
                                    except Exception as exc:
                                        messages[k][kkey]="Errore. "+str(exc)
                                        if updateType=="N":
                                            product.delete_eav(marketplace)
                                        stroke=True
                                        break
                            elif key=="eav_url":
                                for kkey,vvalue in value.items():
                                    if "image" in kkey[0:4]:
                                        file=None
                                        try:
                                            file = urllib.request.urlopen(vvalue).read()
                                        except Exception as exc:
                                            messages[k][kkey]="Errore. "+str(exc)
                                            if updateType=="N":
                                                product.delete_eav(marketplace)
                                            stroke=True
                                            break
                                        try:
                                            image = Image.open(BytesIO(file)).convert("RGB")
                                            filepath=os.path.join(settings.PRIVATE_DIR,product.company.vid,marketplace.id,product.sku,kkey)
                                            image.save(filepath,format="JPEG")
                                        except Exception as exc:
                                            messages[k][kkey]="Errore. "+str(exc)
                                            if updateType=="N":
                                                product.delete_eav(marketplace)
                                            stroke=True
                                            break
                                    urlProduct=ProductUrlEav(sku=product_data["sku"],company=company,attribute=kkey,value=vvalue,marketplace=marketplace)
                                    try:
                                        urlProduct.save()
                                        product.url_eav.add(urlProduct)
                                    except Exception as exc:
                                        messages[k][kkey]="Errore. "+str(exc)
                                        if updateType=="N":
                                            product.delete_eav(marketplace)
                                        stroke=True
                                        break
                            elif key=="eav_char":
                                for kkey,vvalue in value.items():
                                    charProduct=ProductCharEav(sku=product_data["sku"],company=company,attribute=kkey,value=vvalue,marketplace=marketplace)
                                    try:
                                        charProduct.save()
                                        product.char_eav.add(charProduct)
                                    except Exception as exc:
                                        messages[k][kkey]="Errore. "+str(exc)
                                        if updateType=="N":
                                            product.delete_eav(marketplace)
                                        stroke=True
                                        break
                
                if product.int_eav.count()==0 and product.char_eav.count()==0 and product.boolean_eav.count()==0 and product.text_eav.count()==0\
                    and product.url_eav.count()==0 and product.decimal_eav.count()==0:
                    messages[k]["Riga"]="Niente di fatto. Nessuna informazione per questo prodotto"
                    product.delete()
                else:
                    print("creo il prodotto")
                    product.save()
                    messages[k]["Riga"]="Creato"
        except:
            return messages
        return messages

    
        

        #         if cell.value is not None:
        #             ws_row.append(str(cell.value))
        #             stop=False
        #         else:
        #             ws_row.append("")
        #     if stop:
        #         return wslist
        #     else:
        #         wslist.append(ws_row)
        # return wslist






         # def getWSlistCSV(url,delimiter,cut=False):
    #     wslist=[]
    #     d=None
    #     if delimiter=="PV":
    #         d=";"
    #     elif delimiter=="V":
    #         d=","
    #     elif delimiter=="T":
    #         d="\t"
    #     else:
    #         return None

    #     with closing(requests.get(url, stream=True)) as csvfile:
    #         book = csv.reader(codecs.iterdecode(csvfile.iter_lines(),'utf-8-sig'),delimiter=d)
    #         for row in book:
    #             stop=True
    #             ws_row=[]
    #             for cell in row:
    #                 if cell is not None:
    #                     if type(cell) is str and cut:
    #                         ws_row.append(cell[:50])
    #                     else:
    #                         ws_row.append(cell)
    #                     stop=False
    #                 else:
    #                     ws_row.append("")

    #             if stop:
    #                 book.close()
    #                 csvfile.close()
    #                 return wslist
    #             else:
    #                 wslist.append(ws_row)

    #     return wslist