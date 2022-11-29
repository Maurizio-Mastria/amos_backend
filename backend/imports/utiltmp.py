from openpyxl import load_workbook,Workbook
import csv
from marketplaces.models import Marketplace
from products.models import ProductSimple
from products.models import ProductBooleanEav,ProductCharEav,ProductIntEav,ProductDecimalEav,ProductTextEav,ProductUrlEav
from django.db.utils import DataError
from django.conf import settings
import os
from io import BytesIO
from PIL import Image
import urllib
from orders.models import Order,OrderDetail
from customers.models import Customer
from shippings.models import Shipping
from datetime import datetime
 
        
class WorkBook():
    
    def load_workbook_from_path(self,path,isCSV=False):
        self.book = Workbook()
        if isCSV:
            csv_data=[]
            with open(path,"r",encoding="UTF-8") as fp:
                reader=csv.reader(fp,delimiter="\t")
                for row in reader:
                    csv_data.append(row)
            sheet = self.book.active
            for row in csv_data:
                sheet.append(row)
        else:
            self.book=load_workbook(path)

    # def load_workbook_from_url(self,url):
    #     file = urllib.request.urlopen(url).read()
    #     self.book=load_workbook(filename=file)

    # def load_workbook_from_stream(self,filestream):
    #     self.book=load_workbook(filename=filestream)
    #     filestream.close()
        
    def sheetnames(self):
        return self.book.sheetnames

    def close(self):
        self.book.close()

    def datasheet(self,sheetindex,cut=False,limit=None):
        ws=self.book.worksheets[sheetindex]
        wslist=[]
        i=0
        for row in ws.iter_rows():
            if limit is not None and i>limit:
                break
            i+=1
            stop=True
            ws_row=[]
            for cell in row:
                if cell.value is not None:
                    if cut:
                        ws_row.append(str(cell.value)[:80])
                    else:
                        ws_row.append(str(cell.value))
                    stop=False
                else:
                    ws_row.append("")
            if stop:
                return wslist
            else:
                wslist.append(ws_row)
        return wslist


    def importDatasheetToProducts(self,company,sheetindex,jump,columns,rows,marketplaces):
        ws=self.book.worksheets[sheetindex]
        wslist=[]
        messages={}
        updateType="N" # N,U,Overwrite
        j=0
        k=0
        try:
            for row in ws.iter_rows():
                k+=1
                messages[k]={}
                #salto le righe che non servono
                if jump>=0:
                    jump-=1
                    messages[k]["Riga"]="Saltata"
                    continue
                
                j+=1
                try:
                    if rows[j-1] is not True:
                        messages[k]["Riga"]="Riga da non elaborare"
                        continue
                except IndexError:
                    messages[k]["Riga"]="Saltata. Riga non valida"
                    continue
                
                
                i=0
                product_data={}
                product_data["eav_int"]={}
                product_data["eav_boolean"]={}
                product_data["eav_char"]={}
                product_data["eav_text"]={}
                product_data["eav_url"]={}
                product_data["eav_decimal"]={}
                for cell in row:
                    if str(i) in columns:
                        if(columns[str(i)]["name"] in ["sku","gtin","gtin_type"]):
                            product_data[columns[str(i)]["name"]]=cell.value
                        else:
                            if columns[str(i)]["type"]=="INT":
                                product_data["eav_int"][columns[str(i)]["name"]]=int(cell.value)
                            elif columns[str(i)]["type"]=="TEXT":
                                product_data["eav_text"][columns[str(i)]["name"]]=cell.value
                            elif columns[str(i)]["type"]=="BOOLEAN":
                                product_data["eav_boolean"][columns[str(i)]["name"]]=bool(cell.value)
                            elif columns[str(i)]["type"]=="DECIMAL":
                                product_data["eav_decimal"][columns[str(i)]["name"]]=cell.value
                            elif columns[str(i)]["type"]=="URL":
                                product_data["eav_url"][columns[str(i)]["name"]]=cell.value
                            elif columns[str(i)]["type"]=="CHAR":
                                product_data["eav_char"][columns[str(i)]["name"]]=str(cell.value)
                    i+=1
                try:
                    if product_data["sku"] in ["",None]:
                        messages[k]["Riga"]="Saltata. Devi inserire lo SKU"
                        continue
                except KeyError:
                    messages[k]["Riga"]="Saltata. Devi indicare la colonna degli SKU"
                    continue
                stroke=False
                product=None
                if updateType=="N": # Only new products
                    if ProductSimple.objects.filter(company=company,sku=product_data["sku"]).exists():
                        messages[k]["Riga"]="Saltata. Lo SKU esiste e la modalità di import è NUOVO INSERIMENTO"
                        continue
                    product=ProductSimple(company=company,sku=product_data["sku"])
                    try:
                        product.save()
                    except DataError as exc:
                        messages[k]["Riga"]="Saltata. "+str(exc)
                        continue
                elif updateType=="U": # Only existing products
                    try:
                        product=ProductSimple.objects.get(company=company,sku=product_data["sku"])
                    except:
                        messages[k]["Riga"]="Saltata. Lo SKU non esiste"
                        continue
                elif updateType=="O":
                    if ProductSimple.objects.filter(company=company,sku=product_data["sku"]).exists():
                        product=ProductSimple.objects.get(company=company,sku=product_data["sku"])
                    else:
                        product=ProductSimple(company=company,sku=product_data["sku"])
                        try:
                            product.save()
                        except DataError as exc:
                            messages[k]["Riga"]="Saltata. "+str(exc)
                            continue
                for key,value in product_data.items():
                    if(stroke):
                        break
                    if key == "gtin":
                        product.gtin=value
                    elif key == "gtin_type":
                        product.gtin_type=value
                    else:
                        
                        for marketplace in marketplaces.all():
                            if(stroke):
                                break
                            if key=="eav_int":
                                for kkey,vvalue in value.items():
                                    
                                    intProduct=ProductIntEav(sku=product_data["sku"],company=company,attribute=kkey,value=vvalue,marketplace=marketplace)
                                    try:
                                        intProduct.save()
                                        product.add(intProduct)
                                    except Exception as exc:
                                        messages[k][kkey]="Errore. "+str(exc)
                                        if updateType=="N":
                                            product.delete_eav(marketplace)
                                        stroke=True
                                        break
                            elif key=="eav_decimal":
                                for kkey,vvalue in value.items():
                                    decimalProduct=ProductDecimalEav(sku=product_data["sku"],company=company,attribute=kkey,value=vvalue,marketplace=marketplace)
                                    try:
                                        decimalProduct.save()
                                        product.decimal_eav.add(decimalProduct)
                                    except Exception as exc:
                                        messages[k][kkey]="Errore. "+str(exc)
                                        if updateType=="N":
                                            product.delete_eav(marketplace)
                                        stroke=True
                                        break
                            elif key=="eav_boolean":
                                for kkey,vvalue in value.items():
                                    booleanProduct=ProductBooleanEav(sku=product_data["sku"],company=company,attribute=kkey,value=vvalue,marketplace=marketplace)
                                    try:
                                        booleanProduct.save()
                                        product.boolean_eav.add(booleanProduct)
                                    except Exception as exc:
                                        messages[k][kkey]="Errore. "+str(exc)
                                        if updateType=="N":
                                            product.delete_eav(marketplace)
                                        stroke=True
                                        break
                            elif key=="eav_text":
                                for kkey,vvalue in value.items():
                                    textProduct=ProductTextEav(sku=product_data["sku"],company=company,attribute=kkey,value=vvalue,marketplace=marketplace)
                                    try:
                                        textProduct.save()
                                        product.text_eav.add(textProduct)
                                    except Exception as exc:
                                        messages[k][kkey]="Errore. "+str(exc)
                                        if updateType=="N":
                                            product.delete_eav(marketplace)
                                        stroke=True
                                        break
                            elif key=="eav_url":
                                for kkey,vvalue in value.items():
                                    if "image" in kkey[0:4]:
                                        file=None
                                        try:
                                            file = urllib.request.urlopen(vvalue).read()
                                        except Exception as exc:
                                            messages[k][kkey]="Errore. "+str(exc)
                                            if updateType=="N":
                                                product.delete_eav(marketplace)
                                            stroke=True
                                            break
                                        try:
                                            image = Image.open(BytesIO(file)).convert("RGB")
                                            filepath=os.path.join(settings.PRIVATE_DIR,product.company.vid,marketplace.id,product.sku,kkey)
                                            image.save(filepath,format="JPEG")
                                        except Exception as exc:
                                            messages[k][kkey]="Errore. "+str(exc)
                                            if updateType=="N":
                                                product.delete_eav(marketplace)
                                            stroke=True
                                            break
                                    urlProduct=ProductUrlEav(sku=product_data["sku"],company=company,attribute=kkey,value=vvalue,marketplace=marketplace)
                                    try:
                                        urlProduct.save()
                                        product.url_eav.add(urlProduct)
                                    except Exception as exc:
                                        messages[k][kkey]="Errore. "+str(exc)
                                        if updateType=="N":
                                            product.delete_eav(marketplace)
                                        stroke=True
                                        break
                            elif key=="eav_char":
                                for kkey,vvalue in value.items():
                                    charProduct=ProductCharEav(sku=product_data["sku"],company=company,attribute=kkey,value=vvalue,marketplace=marketplace)
                                    try:
                                        charProduct.save()
                                        product.char_eav.add(charProduct)
                                    except Exception as exc:
                                        messages[k][kkey]="Errore. "+str(exc)
                                        if updateType=="N":
                                            product.delete_eav(marketplace)
                                        stroke=True
                                        break
                
                if product.int_eav.count()==0 and product.char_eav.count()==0 and product.boolean_eav.count()==0 and product.text_eav.count()==0\
                    and product.url_eav.count()==0 and product.decimal_eav.count()==0:
                    messages[k]["Riga"]="Niente di fatto. Nessuna informazione per questo prodotto"
                    product.delete()
                else:
                    print("creo il prodotto")
                    product.save()
                    messages[k]["Riga"]="Creato"
        except:
            return messages
        return messages

    
    def importDatasheetToOrders(self,company,sheetindex,jump,columns,rows):
        ws=self.book.worksheets[sheetindex]
        
        messages={}
        j=0
        k=0
        orders={}
        try:
            for row in ws.iter_rows():
                
                
                #salto le righe che non servono
                if jump>=0:
                    jump-=1
                    continue
                
                j+=1
                try:
                    if rows[j-1] is not True:
                        continue
                except IndexError:
                    continue
                
                
                i=0
                order_data={}
                for cell in row:
                    if str(i) in columns:
                        order_data[columns[str(i)]["name"]]=cell.value
                    i+=1
                try:
                    if order_data["order_id"] in ["",None]:
                        continue
                except KeyError:
                    continue
                if order_data["order_id"] not in orders:
                    orders[order_data["order_id"]]=[]
                orders[order_data["order_id"]].append(order_data)

        except:
            print("1")
            messages[0]={"Sono presenti degli errori! Nessun ordine è stato inserito!"}
            return messages

        k=0
        for order_id,details in orders.items():
            print("into")
            k+=1
            messages[k]={}
            

            if Order.objects.filter(company=company,order_id=order_id).exists():
                messages[k][order_id]="L'ordine esiste già!"
                continue
            

            customerObj=None
            shippingObj=None
            orderObj=None
            
            order_price=0
            order_iva=0
            shipping_price=0
            shipping_iva=0
            order_total=0
            shipping_total=0
            order_shipping_total=0
            first=True
            for detail in details:
                print("detail")
                
                marketplace=None
                if detail["marketplace"][:-3].lower()=="amazon":
                    code="AMZ"
                    country=detail["marketplace"][-2:].upper()
                    marketplace=Marketplace.objects.get(company=company,code=code,country=country)
                else:
                    continue
                
                print("superatomarketplace")
                if first:
                    try:
                        customerObj=Customer(company=company,marketplace=marketplace,customer_email=detail["customer_email"],customer_name=detail["customer_name"])
                        print("creato customer")
                    except:
                        messages[k][order_id]="Il cliente non ha un nome e/o un'email validi"
                        first=False
                        break
                    try:
                        customerObj.customer_phone=detail["customer_phone"]
                    except KeyError:
                        customerObj.customer_phone=None
                    try:
                        shippingObj=Shipping(company=company,marketplace=marketplace,shipping_name=detail["shipping_name"],shipping_address=detail["shipping_address"],shipping_city=detail["shipping_city"])
                        shippingObj.shipping_country=detail["shipping_country"]
                        shippingObj.shipping_cap=detail["shipping_cap"]
                        print("creato shipping")
                        if detail["shipping_country"]=="IT":
                            shippingObj.shipping_phone=detail["shipping_phone"]
                            shippingObj.shipping_state=detail["shipping_state"]
                    except:
                        first=False
                        break
                    try:
                        customerObj.save()
                        print("salvato customer")
                    except:
                        first=False
                        break
                    try:
                        shippingObj.save()
                        print("salvato shipping")
                    except DataError as exc:
                        first=False
                        customerObj.delete()
                        messages[k][order_id]=str(exc)
                        break
                    first=False
                try:
                    orderDetailObj=OrderDetail(company=company,marketplace=marketplace,sku=detail["sku"],order_id=detail["order_id"])
                    orderDetailObj.order_item_id=detail["order_item_id"]
                    orderDetailObj.qty=detail["qty"]
                    orderDetailObj.price=detail["price"]
                    orderDetailObj.iva=detail["iva"]
                    orderDetailObj.shipping_price=detail["shipping_price"]
                    orderDetailObj.shipping_iva=detail["shipping_iva"]
                    print("creato detail")
                    
                    # orderDetailObj.title=detail["title"]
                    order_price=float(order_price)+float(detail["price"])
                    order_iva=float(order_iva)+float(detail["iva"])
                    shipping_price=float(shipping_price)+float(detail["shipping_price"])
                    shipping_iva=float(shipping_iva)+float(detail["shipping_iva"])
                    order_total+=order_price+order_iva
                    shipping_total+=shipping_price+shipping_iva
                    order_shipping_total=order_total+shipping_total
                    print("creato detail prezzi")
                except:
                    print("eccezione cancello tutto detail")
                    customerObj.delete()
                    shippingObj.delete()
                    shippingObj=None
                    customerObj=None
                try:
                    if detail["payments_date"] not in [None,""]:
                        orderDetailObj.payments_date=detail["payments_date"]
                        orderDetailObj.status="P"
                except:
                    #default Pending Payment if no date payments is present
                    pass
                try:
                    orderDetailObj.date=detail["date"]
                except:
                    orderDetailObj.date=datetime.now()
                try:                
                    orderDetailObj.business=detail["business"].title()
                except:
                    orderDetailObj.business=False
                try:
                    if "title" in detail and detail["title"] not in [None,""]:
                        orderDetailObj.title=detail["title"]
                    else:
                        title=ProductCharEav.objects.get(company=company,marketplace=marketplace,sku=detail["sku"]).title
                        orderDetailObj.title=title
                except:
                    pass
                
                
                
                if orderObj:
                    print("in order obj")
                    try:
                        orderDetailObj.shipping=shippingObj
                        orderDetailObj.customer=customerObj
                        orderDetailObj.save()
                        print("salvato detail agganciati customer e shipping")
                        orderObj.order_price=order_price
                        orderObj.order_iva=order_iva
                        orderObj.shipping_price=shipping_price
                        orderObj.shipping_iva=shipping_iva
                        orderObj.order_total=order_total
                        orderObj.shipping_total=shipping_total
                        orderObj.order_shipping_total=order_shipping_total
                        orderObj.save()
                        print("salvato order")
                        orderObj.order_detail.add(orderDetailObj)
                        print("salvato order agganciato detail")
                    except:
                        print("delete order e detail")
                        for orderDetail in orderObj.order_detail.all():
                            orderDetail.delete()
                        orderObj.delete()
                        customerObj.delete()
                        shippingObj.delete()
                else:
                    print("in ordre else")
                    try:
                        orderObj=Order(company=company,marketplace=marketplace,order_id=order_id)
                        print("creato order")
                        orderDetailObj.shipping=shippingObj
                        orderDetailObj.customer=customerObj
                        orderDetailObj.save()
                        print("salvato order detail else")
                        orderObj.order_price=order_price
                        orderObj.order_iva=order_iva
                        orderObj.shipping_price=shipping_price
                        orderObj.shipping_iva=shipping_iva
                        orderObj.order_total=order_total
                        orderObj.shipping_total=shipping_total
                        orderObj.order_shipping_total=order_shipping_total
                        orderObj.save()
                        print("salvato order else")
                        orderObj.order_detail.add(orderDetailObj)
                    except Exception as exc:
                        print("eccezione else order customer shipping")
                        customerObj.delete()
                        shippingObj.delete()

        
        return messages